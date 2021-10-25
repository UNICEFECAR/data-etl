"""
Utility functions that operate directly on our data-dictionary in Excel using pandas
Typically retrieve queries, useful for our ETL purposes
To improve: implement extraction as an abstract class following James' refactoring
"""

from os import name
import requests
import pandas as pd
import pandasdmx as pdsdmx
import numpy as np
import re
import warnings
from pandas_datareader import wb
from bs4 import BeautifulSoup
from io import BytesIO, StringIO
from difflib import SequenceMatcher
from heapq import nlargest
from time import sleep
from zipfile import ZipFile


def get_API_code_address_etc(excel_data_dict):
    """ filters all indicators that are extracted by API
        :param excel_data_dict: path/file to our excel data dictionary in repo
        :return: pandas dataframe (df) with code, url endpoint for requests, data source name, etc
    """
    # read snapshots table from excel data-dictionary
    snapshot_df = pd.read_excel(excel_data_dict, sheet_name="Snapshot")
    # read sources table from excel data-dictionary
    sources_df = pd.read_excel(excel_data_dict, sheet_name="Source")
    # join snapshot and source based on Source_Id
    # 'left' preserves key order from snapshots
    snap_source_df = snapshot_df.merge(
        sources_df, on="Source_Id", how="left", sort=False
    )
    # read indicators table from excel data-dictionary
    indicators_df = pd.read_excel(excel_data_dict, sheet_name="Indicator")
    # join snap_source and indicators based on Indicator_Id
    snap_source_ind_df = snap_source_df.merge(
        indicators_df, on="Indicator_Id", how="left", sort=False
    )
    # read Value_type table from excel data-dictionary
    val_df = pd.read_excel(excel_data_dict, sheet_name="Value_type")
    # join snap_source_ind and Value_type based on Value_Id
    snap_source_ind_val_df = snap_source_ind_df.merge(
        val_df, on="Value_Id", how="left", sort=False
    )
    # read Transformation table from excel data-dictionary
    transf_df = pd.read_excel(excel_data_dict, sheet_name="Transformation")
    # join snap_transform_ind and Tranformation based on Transformation_Id
    snap_join_df = snap_source_ind_val_df.merge(
        transf_df, on="Transformation_Id", how="left", sort=False
    )
    # get indicator codes, url endpoints, etc for API/Web Scrape/Calculation extractions
    logic_API = snap_join_df.Type_x.str.contains("API|Web Scrape|Calculation")
    logic_not_null = logic_API.notnull()
    # operator AND for two logics above
    logic_API_not_null = logic_API & logic_not_null

    api_code_addr_etc_df = snap_join_df[logic_API_not_null][
        [
            "Type_x",
            "Theme",
            "Code_y",
            "Name",
            "Address",
            "Name_y",
            "Comments_y",
            "Content_type",
            "Units_y",
            "Disaggregation",
            "Freq_Coll",
            "Nature",
            "Unit_Mult",
            "Parameters",
            "Numerator",
            "Denominator",
        ]
    ]

    api_code_addr_etc_df.rename(
        columns={
            "Type_x": "Extraction_type",
            "Name": "Indicator_name",
            "Code_y": "Code",
            "Name_y": "Data_Source",
            "Comments_y": "Obs_Footnote",
            "Units_y": "Units",
            "Parameters": "Transf_Param",
        },
        inplace=True,
    )
    # return dataframe (note its df index will correspond to that of snapshots)
    return api_code_addr_etc_df


# function for api request as proposed by Daniele
# errors are printed and don't stop program execution
def api_request(address, params=None, headers=None, timeout=45):
    """
    TODO: Look at the error handerling here
    """
    try:
        response = requests.get(
            address, params=params, headers=headers, timeout=timeout
        )
        # If the response was successful, no Exception will be raised
        response.raise_for_status()
    except requests.exceptions.HTTPError as http_err:
        print(f"HTTP error occurred: {http_err}")
        # return None
    except Exception as error:
        print(f"Other error occurred: {error}")
        # return None
    # return response object (check with James notes below)
    # response will not be defined for the following errors:
    # "no connection adapters found", "Invalid URL: No schema supplied"
    # same response not defined for timeouts: alternative to return response None in those cases
    # TODO: evaluate the impact of the current flow managing response return None for errors
    return response


def data_reader(address, country_codes=None, start_period=None, end_period=None):
    """
    Uses pandas data reader to download World Bank indicators (data and metadata)
    :param address: from data dictionary, string containing World Bank Indicator code/s
    :param country_codes: TMEE countries to call (note data reader default just 'US', 'CA' and 'MX')
    :param start_period: First year of the data series (note data reader default is 2003)
    :param end_period: Last year of the data series, inclusive (note data reader default is 2005)
    :return: pandas dataframe with data/metadata and error flag
    World Bank sex disaggregation is compilated using different indicator codes with 'FE' and 'MA'
    Dev note: address must contain either one indicator (no sex) or three indicators (total, 'FE' and 'MA')
    TODO: handle age disaggregation from World Bank codes ! Nice add-on feature
    Transformations must be done to flaten data reader output
    Simple error handling prints messages and don't stop program execution
    Error message: when any of the provided indicator codes is wrong
    """

    # initialize output
    data_df = None
    data_error = False

    # get indicator code/s from data dictionary
    wb_codes = re.findall(r"'(.+?)'", address)

    # add country_call if present
    country_call = None
    if country_codes:
        country_call = list(country_codes.values())

    # pandas data reader data call
    # handle warnings as errors forces all indicator codes to be correct!
    warnings.filterwarnings("error")
    try:
        # pandas data reader country calls (ISO-alpha 3)
        data_df = wb.download(
            indicator=wb_codes, country=country_call, start=start_period, end=end_period
        )
    except Exception as error:
        print(f"Error occurred: {error}")
        data_error = True

    if not data_error:

        # flatten output
        data_df.reset_index(inplace=True)

        # set warnings back to default for metadata call
        warnings.resetwarnings()
        warnings.filterwarnings("ignore", category=ResourceWarning)
        warnings.filterwarnings("ignore", category=DeprecationWarning)
        # call for metadata using first element in wb_codes list
        metadata_df = wb.search(string=wb_codes[0], field="id")
        # metadata will download and cache all indicators the first time is called
        # depending on your connection this can take some time
        # subsequent searches should be faster on the chached copy

        # filter metadata_df (wb.search does fuzzy matching)
        filter_id = metadata_df.id == wb_codes[0]
        metadata_row = metadata_df[filter_id]

        # data source: collection
        collection = metadata_row.source.values[0].strip(" .")
        # data source: topic
        topic = metadata_row.topics.values[0].strip(" .")
        # data source: organization
        organization = metadata_row.sourceOrganization.values[0].decode().strip(" .")

        # acknowledge World Bank data compilation
        wb_collection = f"Compiled in World Bank collection: {collection} ({topic})."
        # full metadata to publish in data_source
        data_source = f"{organization}. {wb_collection}"
        data_df["source"] = data_source

        # add country codes from pandas data reader and avoid posterior country mappings
        country_df = wb.get_countries()
        country_df.rename(columns={"name": "country"}, inplace=True)
        # add iso3 codes to data_df from country_df
        data_df = data_df.merge(
            country_df[["country", "iso3c"]], on="country", how="left", sort=False
        )

        # boolean sex_female array
        sex_female = np.array(["FE" in elem.split(".") for elem in wb_codes])
        # boolean sex_male array
        sex_male = np.array(["MA" in elem.split(".") for elem in wb_codes])
        # boolean sex_total array
        sex_total = ~(sex_female | sex_male)

        # accepted disaggregation: total or (total and female and male)
        if len(wb_codes) == 1 and sex_total.any():
            # add sex dimension all populated with total
            data_df["sex"] = "total"
            # rename column with indicator name as value
            data_df.rename(columns={wb_codes[0]: "value"}, inplace=True)

        elif (
            len(wb_codes) == 3
            and sex_total.any()
            and sex_female.any()
            and sex_male.any()
        ):
            # rename columns using sex disaggregation
            data_df.rename(
                columns={wb_codes[np.where(sex_total)[0][0]]: "value_total"},
                inplace=True,
            )
            data_df.rename(
                columns={wb_codes[np.where(sex_female)[0][0]]: "value_female"},
                inplace=True,
            )
            data_df.rename(
                columns={wb_codes[np.where(sex_male)[0][0]]: "value_male"}, inplace=True
            )

            # wide to long transformation: SDMX has sex as a dimension
            data_df_long = pd.wide_to_long(
                data_df,
                stubnames="value",
                i=["country", "year"],
                j="sex",
                sep="_",
                suffix=r"\w+",
            )

            # overwrite output variable with long format flatten
            data_df = data_df_long.reset_index()

        else:
            # flag data_error to avoid posterior mapping of raw data
            data_error = True
            print("Verify correct sex disaggregation in pandas data reader call")

    return data_df, data_error


def web_scrape(raw_html, source_key=None):
    """
    Web scraping (data and metadata)
    :param raw_html: this is a satisfactory url requests response (HTML content)
    :param source_key: future dev of different web structures
    Dev note: web structure now is WHO: https://apps.who.int/immunization_monitoring/globalsummary/
    :return: pandas dataframe with data
    Simple wide to long transformation to match SDMX dimensions
    TODO: error handling?
    TODO: generalize metadata extraction?
    """
    # correct WHO HTML (thank you developer!)
    html_text = StringIO(raw_html.text)
    # Erase % from colspan (possible regex equivalence?)
    html_refactor = "".join(
        [line.replace("%", "") if "colspan" in line else line for line in html_text]
    )
    # Soupify refactored html
    soup = BeautifulSoup(html_refactor, "html.parser")
    # Extract tables (missing targeting attributes)
    tables = soup.find_all("table")
    # Feed target table by HTML position into pandas: specify [0] to get actual df
    data_df = pd.read_html(str(tables[2]), header=0, skiprows=[1])[0]
    # drop columns by regex
    data_df.drop(
        list(
            data_df.filter(
                regex=re.compile("unnamed|category|indicator", re.IGNORECASE)
            )
        ),
        axis=1,
        inplace=True,
    )
    # rename first column as country
    data_df.rename(columns={data_df.columns[0]: "country"}, inplace=True)

    # transform wide to long (except last 2 rows)
    data_df = pd.melt(data_df.iloc[:-2], id_vars=["country"])
    data_df.rename(columns={"variable": "year"}, inplace=True)

    # dummy use of source key to trigger metadata extraction
    if source_key:
        # metadata (data provider): get it from last row (this should be generalized)
        data_prov = data_df.iat[-1, 0].strip("")
        # metadata (last updated): target table by HTML position - contains also class attribute -
        data_date = pd.read_html(str(tables[1]))[0].iat[-1, 0].strip("")
        data_date = data_date.replace("Next", ". Next") + "."

        # full metadata to publish in data_source (this should be generalized for posterior structure mapping)
        data_source = f"{data_prov}{data_date}"
        data_df["source"] = data_source

    # country names to lowercase for code mapping
    data_df["country"] = data_df.country.str.lower()

    return data_df


def estat_reader(address, raw_path, ind_code, params=None, headers=None):
    """
    First wrapper for sdmx eurostat
    :param address: expected to contain endpoint and query
    :param raw_path: path to write xml raw file
    :param ind_code: name to write xml raw file
    :param params: requests parameters
    :param headers: requests headers
    :return: download flag: True/False
    """

    # get dataflow from address
    url_split = address.split("/")
    # get position of dataflow
    dflow_position = url_split.index("data") + 1
    dflow_name = url_split[dflow_position]
    dsd_name = f"DSD_{dflow_name}"

    # use pdsdmx to get dsd from estat
    Estat = pdsdmx.Request("ESTAT", backend="memory")

    # handle errors from pdsdmx dsd request
    try:
        print(
            f"Querying EUROSTAT metadata: please wait, long response times are reported"
        )
        Dsd_estat = Estat.datastructure(dsd_name)
        # If the response was successful, no Exception will be raised
        Dsd_estat.response.raise_for_status()
    except requests.exceptions.HTTPError as http_err:
        print(f"HTTP error occurred: {http_err}")
        return False
    except Exception as error:
        print(f"Other error occurred: {error}")
        return False

    flag_download = False
    # first api call (determine if xml response asynchronous)
    print(f"Querying EUROSTAT data: please wait, long response times are reported")
    indicator_xml = api_request(address, params, headers, timeout=15)

    if hasattr(indicator_xml, "status_code"):
        if indicator_xml.status_code == 200:
            # write raw xml file first: can't make pandasdmx work from requests content
            raw_file = f"{raw_path}{ind_code}.xml"
            with open(raw_file, "wb") as f:
                f.write(indicator_xml.content)
            # now pandasdmx read written file
            xml_message = pdsdmx.read_sdmx(
                raw_file, format="XML", dsd=Dsd_estat.structure[dsd_name]
            )

        # inspect asynchronous response code: 413 (actually not None)
        if hasattr(xml_message.footer, "text"):
            print(f"EUROSTAT asynchronous response for {ind_code}")
            # eurostat data in file url/zip file
            file_url = str(xml_message.footer.text[0]).split("URL:")[1].strip()

            # try attemps to get file in url with wait_time duplicating at each attempt
            wait_time, attempts = (15, 5)
            for a in range(attempts):
                wait_time = wait_time * 2
                print(
                    f"Attempt {a+1} to reach EUROSTAT zip file for {ind_code}, please wait: {wait_time} seconds"
                )
                sleep(wait_time)
                indicator_xml = api_request(file_url)
                # status code 200: ZIP response exists
                if hasattr(indicator_xml, "status_code"):
                    if indicator_xml.status_code == 200:
                        flag_download = True
                        break

            if flag_download:
                # Open the zip archive
                zf = ZipFile(BytesIO(indicator_xml.content), mode="r")
                # The archive should contain only one file
                file_in_zip = zf.infolist()[0]

                # overwrite xml with zf
                with open(raw_file, "wb") as f:
                    f.write(zf.open(file_in_zip).read())

                # proper read_sdmx: content inside zip
                xml_message = pdsdmx.read_sdmx(
                    raw_file, format="XML", dsd=Dsd_estat.structure[dsd_name]
                )

        else:
            # no footer assumes data delivered at first time
            flag_download = True

        if flag_download:
            print(f"Parsing {ind_code} sdmx-xml to pandas: please wait")
            # convert xml to pandas (long format, attributes per observation)
            data_df = xml_message.to_pandas(dtype=str, attributes="o", rtype="rows")
            data_df.reset_index(inplace=True)
            # EUROSTAT reports all dimension keys even if empty values: detect/drop strings 'NaN'
            filter_nan_str = pd.to_numeric(data_df.value, errors="coerce").isnull()
            data_df.drop(data_df[filter_nan_str].index, inplace=True)
            # save raw file as csv
            data_df.to_csv(f"{raw_path}{ind_code}.csv", index=False)
            print(f"Indicator {ind_code} succesfully downloaded")

    return flag_download


def oecd_reader(address, raw_path, ind_code, params=None, headers=None):
    """
    First wrapper for sdmx oecd
    :param address: expected to contain endpoint and query
    :param raw_path: path to write json and csv raw files
    :param ind_code: name to write json and csv raw file
    :param params: requests parameters
    :param headers: requests headers
    :return: download flag: True/False
    """

    flag_download = False
    # api call for json file
    indicator_json = api_request(address, params, headers)

    if indicator_json.status_code == 200:
        # write raw json file first: can't make pandasdmx work from requests content
        raw_file = f"{raw_path}{ind_code}.json"
        with open(raw_file, "wb") as f:
            f.write(indicator_json.content)
        # now pandasdmx read written file
        json_message = pdsdmx.read_sdmx(raw_file, format="JSON")

        print(f"Parsing {ind_code} sdmx-json to pandas: please wait")
        # convert json to pandas (long format, attributes per observation)
        data_df = json_message.to_pandas(dtype=str, attributes="o", rtype="rows")

        # it's observed that names in multi-index may repeat in df columns, then check and rename
        check_idx_col_names = [item in data_df.columns for item in data_df.index.names]
        if any(check_idx_col_names):
            # rename first
            rename_array = [
                (name + "_idx") if check_idx_col_names[i] else name
                for i, name in enumerate(data_df.index.names)
            ]
            data_df.index.rename(rename_array, inplace=True)

        data_df.reset_index(inplace=True)

        # save raw file as csv
        data_df.to_csv(f"{raw_path}{ind_code}.csv", index=False)
        print(f"Indicator {ind_code} succesfully downloaded")
        flag_download = True

    return flag_download


def undp_reader(address, raw_path, ind_code, params=None, headers=None):
    """
    First wrapper for undp api
    :param address: expected to contain endpoint and query
    :param raw_path: path to write csv raw files
    :param ind_code: name to write csv raw file
    :param params: requests parameters
    :param headers: requests headers
    :return: download flag: True/False
    """

    flag_download = False
    # api call for json file
    indicator_json = api_request(address, params, headers)

    if indicator_json.status_code == 200:
        # transform json response into csv file using pandas
        # step one - read json and normalize
        wide_df = pd.json_normalize(
            indicator_json.json()["indicator_value"], max_level=3
        )
        # step two - build long format from column names structure
        long_df = wide_df.columns.str.split(".", expand=True).to_frame(
            index=False, name=["cou", "ind", "year"]
        )
        # step three fill values from series
        long_df["val"] = wide_df.values[0]

        # save raw file as csv
        long_df.to_csv(f"{raw_path}{ind_code}.csv", index=False)
        print(f"Indicator {ind_code} succesfully downloaded")
        flag_download = True

    return flag_download


def calculate_indicator(func_name, func_param, numerator, denominator, dest_dict):
    """
    First wrapper for calculations
    :param func_name: function name to be called
    :param func_param: input parameters to func_name
    :param numerator: original indicator numerator
    :param denominator: original indicator denominator
    :param dest_dict: info dictionary for calculated indicator
    :return: download flag: True
    TODO: error handling
    """

    orig_indicators = [numerator, denominator]
    globals()[func_name](func_param, orig_indicators, dest_dict)

    return True


def calc_age_sum(age_limits, orig_ind, dest_info):
    """
    First wrapper for group age calculations - Accomodates loop for age disaggregation
    :param age_limits: string with age limits
    :param orig_ind: original indicator codes to read
    :param dest_info: indicator_code and path to write calculated data
    :return: download flag: True
    TODO: error handling
    """

    # configuration of age disaggregation
    age_disagg = age_limits.split(",")

    # read indicator raw data (numerator only)
    org_code = orig_ind[0]
    raw_path = dest_info["path"]
    data_raw = pd.read_csv(f"{raw_path}{org_code}.csv", dtype=str)

    # retain only codes from csv headers
    raw_columns = data_raw.columns.values
    rename_dict = {k: v.split(":")[0] for k, v in zip(raw_columns, raw_columns)}
    data_raw.rename(columns=rename_dict, inplace=True)

    # retain only codes from age groups, indicator and obs_status column
    data_raw.loc[:, "AGE"] = data_raw.AGE.apply(lambda x: x.split(":")[0])
    data_raw.loc[:, "INDICATOR"] = data_raw.INDICATOR.apply(lambda x: x.split(":")[0])
    obs_status_notnan = data_raw.OBS_STATUS.notnull()
    data_raw.loc[obs_status_notnan, "OBS_STATUS"] = data_raw[
        obs_status_notnan
    ].OBS_STATUS.apply(lambda x: x.split(":")[0])

    # age calculations loop
    df_out = None
    for age_group in age_disagg:
        # year start/end: empty year assumes UNPD limit --> [0;100]
        y0_1 = age_group.strip(" ()").split(";")
        y0_1[0] = y0_1[0] if y0_1[0] != "" else "0"
        y0_1[1] = y0_1[1] if y0_1[1] != "" else "100"
        single_ages = [
            "Y0" + str(i) if ((i < 10) & (i != 0)) else "Y" + str(i)
            for i in range(int(y0_1[0]), int(y0_1[1]) + 1)
        ]

        # query data_raw years
        query_y = "AGE in @single_ages"
        df_q = (
            data_raw.query(query_y)
            .astype({"OBS_VALUE": float})
            .groupby(by=["REF_AREA", "SEX", "TIME_PERIOD"], sort=False)
            .agg({"OBS_VALUE": "sum"})
            .round(3)
        ).reset_index()

        # keep unit multiplier as in the original source
        # commented alternative below: bring unit multiplier into obs_value
        # Unit Multiplier must be rewriten to zero (if incorporated into obs_value)

        # u_mult = data_raw.UNIT_MULTIPLIER.unique()[0]
        # u_mult = int(u_mult.split(":")[0]) if not pd.isnull(u_mult) else 0
        # df_q.loc[:, "OBS_VALUE"] = (df_q.OBS_VALUE * 10 ** int(u_mult)).astype(int)

        # add age group accordingly
        df_q["AGE"] = f"Y{y0_1[0]}T{y0_1[1]}" if y0_1[1] != "100" else f"Y_GE{y0_1[0]}"
        # replace Y_GE0 as _T
        df_q.AGE.replace({"Y_GE0": "_T"}, inplace=True)

        # concat output
        df_out = pd.concat([df_out, df_q], ignore_index=True)

    # complete data raw structure
    for col in [col for col in data_raw.columns if len(data_raw[col].unique()) == 1]:
        df_out[col] = data_raw[col].unique()[0]

    # indicator destination code
    dest_code = dest_info["code"]
    # transform indicator column to destination code
    df_out["INDICATOR"].replace(org_code, dest_code, inplace=True)
    # place DATAFLOW column index 0 (ETL requirement: Helix transforms)
    df_out.insert(0, "DATAFLOW", df_out.pop("DATAFLOW"))

    # unpd obs_status projection
    pr_years = data_raw[data_raw.OBS_STATUS == "PR"].TIME_PERIOD.unique()
    pr_rows = [df_out.TIME_PERIOD == yr for yr in pr_years]
    df_out.loc[np.logical_or.reduce(pr_rows), "OBS_STATUS"] = "PR"

    # save data_raw file and return flag
    df_out.to_csv(f"{raw_path}{dest_code}.csv", index=False)
    print(f"Indicator {dest_code} succesfully calculated")

    return True


def calc_indicator_rate(func_param, orig_ind, dest_info):
    """
    First wrapper for rate calculations
    :param func_param: use to query the proper age groups so far
    :param orig_ind: indicator codes to read (numerator, denominator)
    :param dest_info: indicator_code and path to write calculated data
    :return: download flag: True
    TODO: error handling
    """

    # age groups query
    age_groups = func_param.split(",")

    # read numerator raw data
    num_code = orig_ind[0]
    raw_path = dest_info["path"]
    num_raw = pd.read_csv(f"{raw_path}{num_code}.csv", dtype=str)

    # retain only codes from numerator csv headers
    num_raw_col = num_raw.columns.values
    rename_dict = {k: v.split(":")[0] for k, v in zip(num_raw_col, num_raw_col)}
    num_raw.rename(columns=rename_dict, inplace=True)

    # retain only codes from indicator column
    num_raw.loc[:, "INDICATOR"] = num_raw.INDICATOR.apply(lambda x: x.split(":")[0])

    # read denominator raw data
    den_code = orig_ind[1]
    den_raw = pd.read_csv(f"{raw_path}{den_code}.csv", dtype=str)

    # retain only codes from denominator csv headers
    den_raw_col = den_raw.columns.values
    rename_dict = {k: v.split(":")[0] for k, v in zip(den_raw_col, den_raw_col)}
    den_raw.rename(columns=rename_dict, inplace=True)

    # calcultion index
    calc_index = ["REF_AREA", "SEX", "TIME_PERIOD"]

    # age groups to query
    y_num = age_groups[0].strip(" ()")
    y_den = age_groups[1].strip(" ()")

    # share of num against den_tot
    share_pop = (
        (
            num_raw.query("AGE == @y_num")
            .astype({"OBS_VALUE": float})
            .set_index(calc_index)
            .OBS_VALUE
            / den_raw.query("AGE == @y_den")
            .astype({"OBS_VALUE": float})
            .set_index(calc_index)
            .OBS_VALUE
            * 100
        )
        .round(2)
        .reset_index()
    )

    # complete remaining columns: join extraction source data structure
    share_pop = share_pop.merge(
        num_raw.drop(columns=["OBS_VALUE", "UNIT_MEASURE", "UNIT_MULTIPLIER"]),
        on=calc_index,
        how="left",
        sort=False,
    )

    # indicator destination code
    dest_code = dest_info["code"]
    # transform indicator column to destination code
    share_pop["INDICATOR"].replace(num_code, dest_code, inplace=True)
    # place DATAFLOW column index 0 (ETL requirement: Helix transforms)
    share_pop.insert(0, "DATAFLOW", share_pop.pop("DATAFLOW"))
    # replace unit with data dictionary info
    share_pop["UNIT_MEASURE"] = dest_info["units"]
    # trivial unit multiplier: zero always for a rate
    share_pop["UNIT_MULTIPLIER"] = 0

    # save data_raw file and return flag
    share_pop.to_csv(f"{raw_path}{dest_code}.csv", index=False)
    print(f"Indicator {dest_code} succesfully calculated")

    return True


def get_close_match_indexes(word, possibilities, n=3, cutoff=0.6):
    """Use SequenceMatcher to return a list of indexes of the best "good enough" matches
    :param word: string
    :param possibilities: list of strings
    :param n: (optional integer) maximum number of close matches to return
    :param cutoff: (optional) Threshold score to ignore possibilities
    """
    result = []
    s = SequenceMatcher()
    s.set_seq2(word)
    for idx, x in enumerate(possibilities):
        s.set_seq1(x)
        if s.ratio() >= cutoff:
            result.append((s.ratio(), idx))

    # Move the best scorers to head of list
    result = nlargest(n, result)

    # Strip scores for the best n matches
    return [x for score, x in result]


def get_codelist_API_legacy(data_dict_df, path_legacy):
    """
    wraps all indicators codes and names taken from data dictionary and
    adds to these the indicator codes and names from legacy data
    :param data_dict_df: df from data dictionary query (get_API_code_address_etc function output)
    :path_legacy: path/file to legacy indicators meta data (age, sex, code, units)
    returns df with indicator code and name (typicall SDMX codelist)
    """

    # import csv into pandas
    legacy_meta_data = pd.read_csv(path_legacy, dtype=str)

    # search for sex and age totals
    sex_and_age_t = (legacy_meta_data.age == "_T") & (legacy_meta_data.sex == "_T")

    # TMEE legacy indicator pattern to be deleted (e.g: 4.1.7)
    ind_pattern = r"\d+\.\d+.\d+."

    # make new data frame with legacy code and name only
    legacy_code_name_df = pd.concat(
        [
            legacy_meta_data.code[sex_and_age_t].str.strip(),
            legacy_meta_data.indicator[sex_and_age_t]
            .replace(ind_pattern, "", regex=True)
            .str.strip(),
        ],
        axis=1,
    )
    # rename df columns to match data_dict_df
    legacy_code_name_df.rename(
        columns={"code": "Code", "indicator": "Indicator_name"}, inplace=True
    )

    # search indicators not listed under sex and age totals
    missed_codes = np.setdiff1d(
        legacy_meta_data.code.unique(), legacy_code_name_df.Code.unique()
    )

    # add missed_codes in missed_legacy (with empty indicator names if code is repated)
    missed_legacy = [
        {
            "Code": code,
            "Indicator_name": legacy_meta_data.indicator[legacy_meta_data.code == code]
            .replace(ind_pattern, "", regex=True)
            .str.strip()
            .values[0],
        }
        if ((legacy_meta_data.code == code).sum() == 1)
        else {"Code": code, "Indicator_name": ""}
        for code in missed_codes
    ]

    # return concatenation of data_dict_df and legacy_code_name_df (missed_legacy appended)
    return pd.concat(
        [
            data_dict_df[["Code", "Indicator_name"]],
            legacy_code_name_df.append(missed_legacy),
        ]
    )


def get_units_codelist(etl_out_file):
    """
    wraps all unique unit codes from etl_out_file 'csv' file
    flags if there is any entry without units specified
    :param etl_out_file: path/file to 'csv' with all transformed data
    returns df with units code and empty name/label (typicall SDMX codelist)
    """

    etl_out_df = pd.read_csv(etl_out_file, dtype=str)

    # get unique codes
    unit_codes_array = etl_out_df.UNIT_MEASURE.unique()
    # empty array for code labels (filled with NaN's)
    unit_label_array = np.empty(len(unit_codes_array))
    unit_label_array[:] = np.NaN
    # prepare dataframe output
    cl_units_df = pd.DataFrame(
        {"Code": unit_codes_array, "Unit_label": unit_label_array}
    )

    # flag empty
    empty_flag = etl_out_df.UNIT_MEASURE.isnull().any()

    return cl_units_df, empty_flag


# function from stackoverflow by MaxU
def append_df_to_excel(
    filename,
    df,
    sheet_name="Sheet1",
    startrow=None,
    truncate_sheet=False,
    **to_excel_kwargs,
):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if "engine" in to_excel_kwargs:
        to_excel_kwargs.pop("engine")

    writer = pd.ExcelWriter(filename, engine="openpyxl")

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError
        raise IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    # would this below overwrite startrow defined above? (commented by beto S.)
    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
